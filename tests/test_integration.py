"""Integration tests for full pipeline."""

import pytest
import pandas as pd
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.features.roll_pressure import RollPressureCalculator
from src.viz.heatmap import HeatmapGenerator
from src.viz.excel_alert import ExcelAlertExporter


class TestFullPipeline:
    """Test complete roll pressure pipeline."""

    def test_end_to_end_with_mock_data(self, mock_config, mock_cftc_data, mock_oi_data, mock_roll_pressure_data):
        """Test complete pipeline with mock data."""
        # This tests that all components work together

        calc = RollPressureCalculator(mock_config)

        # Test merge
        merged = calc.merge_data(mock_cftc_data, mock_oi_data)
        assert len(merged) > 0

        # Add mock days to expiry
        merged['days_to_expiry'] = 5

        # Calculate roll pressure
        result = calc.calculate_roll_pressure(merged)
        assert 'roll_pressure' in result.columns

        # Add alerts
        final = calc.add_alert_column(result)
        assert 'ALERTE_48H' in final.columns

    def test_heatmap_generation(self, mock_roll_pressure_data, mock_config, temp_output_dir):
        """Test heatmap generation."""
        # Update config with temp paths
        mock_config['output_files']['heatmap_png'] = str(temp_output_dir / "test_heatmap.png")

        generator = HeatmapGenerator(mock_config)

        # Generate PNG
        png_path = generator.generate_png_heatmap(mock_roll_pressure_data)

        assert Path(png_path).exists()
        assert Path(png_path).stat().st_size > 0

    def test_excel_export(self, mock_roll_pressure_data, mock_config, temp_output_dir):
        """Test Excel export."""
        # Update config with temp path
        mock_config['output_files']['excel'] = str(temp_output_dir / "test_output.xlsx")

        exporter = ExcelAlertExporter(mock_config)

        # Export
        excel_path = exporter.export_to_excel(mock_roll_pressure_data)

        assert Path(excel_path).exists()
        assert Path(excel_path).stat().st_size > 0

        # Verify it's a valid Excel file (openpyxl can open it)
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)

        # Check sheets exist
        assert 'Summary' in wb.sheetnames
        assert 'Data' in wb.sheetnames
        assert 'Heatmap' in wb.sheetnames

    def test_alert_detection(self, mock_roll_pressure_data):
        """Test that alerts are properly detected."""
        # Should have alerts for BRENT when days_to_expiry <= 2
        alerts = mock_roll_pressure_data[mock_roll_pressure_data['ALERTE_48H'] == True]

        assert len(alerts) > 0
        assert all(alerts['days_to_expiry'] <= 2)
        assert all(alerts['roll_pressure'] > 0.65)

    def test_no_alerts_when_pressure_low(self, mock_config):
        """Test no alerts when roll pressure is low."""
        # Create data with low roll pressure
        data = pd.DataFrame({
            'date': pd.date_range('2025-08-01', '2025-08-05', freq='D'),
            'market': ['WTI'] * 5,
            'spec_net_long': [100000] * 5,
            'open_interest': [500000] * 5,
            'days_to_expiry': [5, 4, 3, 2, 1],
            'positioning_ratio': [0.2] * 5,
            'roll_pressure': [1.0, 0.8, 0.6, 0.4, 0.2]  # All below threshold
        })

        calc = RollPressureCalculator(mock_config)
        result = calc.add_alert_column(data)

        # No alerts should be triggered (roll_pressure < 0.65)
        assert result['ALERTE_48H'].sum() == 0

    def test_multiple_markets_processing(self, mock_cftc_data, mock_oi_data, mock_config):
        """Test processing multiple markets simultaneously."""
        calc = RollPressureCalculator(mock_config)

        merged = calc.merge_data(mock_cftc_data, mock_oi_data)

        # Should have both markets
        markets = merged['market'].unique()
        assert 'WTI' in markets
        assert 'BRENT' in markets

        # Each market should have equal number of records
        wti_count = len(merged[merged['market'] == 'WTI'])
        brent_count = len(merged[merged['market'] == 'BRENT'])
        assert wti_count == brent_count


class TestDataQuality:
    """Test data quality and validation."""

    def test_no_negative_roll_pressure(self, mock_roll_pressure_data):
        """Test that roll pressure values are never negative."""
        assert (mock_roll_pressure_data['roll_pressure'] >= 0).all()

    def test_roll_pressure_bounded(self, mock_roll_pressure_data):
        """Test that roll pressure respects bounds."""
        assert (mock_roll_pressure_data['roll_pressure'] <= 2.0).all()

    def test_no_missing_dates_in_range(self, mock_roll_pressure_data):
        """Test that there are no gaps in date coverage."""
        for market in mock_roll_pressure_data['market'].unique():
            market_data = mock_roll_pressure_data[mock_roll_pressure_data['market'] == market]

            dates = market_data['date'].sort_values()
            date_diffs = dates.diff().dropna()

            # All diffs should be 1 day (no gaps)
            assert all(date_diffs == pd.Timedelta(days=1))

    def test_positioning_ratio_calculation(self, mock_roll_pressure_data):
        """Test positioning ratio is correctly calculated."""
        for _, row in mock_roll_pressure_data.iterrows():
            expected_ratio = row['spec_net_long'] / row['open_interest']
            assert abs(row['positioning_ratio'] - expected_ratio) < 0.0001

    def test_roll_pressure_formula(self, mock_roll_pressure_data):
        """Test roll pressure formula is correct."""
        for _, row in mock_roll_pressure_data.iterrows():
            expected_rp = row['positioning_ratio'] * row['days_to_expiry']
            expected_rp = min(expected_rp, 2.0)  # Apply ceiling

            assert abs(row['roll_pressure'] - expected_rp) < 0.0001


class TestEdgeCases:
    """Test edge cases and error handling."""

    def test_empty_dataframe_handling(self, mock_config):
        """Test handling of empty DataFrames."""
        calc = RollPressureCalculator(mock_config)

        empty_df = pd.DataFrame(columns=['date', 'market', 'spec_net_long'])

        # Should not crash
        result = calc.add_alert_column(empty_df)
        assert 'ALERTE_48H' in result.columns

    def test_single_day_data(self, mock_config):
        """Test with single day of data."""
        data = pd.DataFrame({
            'date': [pd.Timestamp('2025-08-01')],
            'market': ['WTI'],
            'spec_net_long': [150000],
            'open_interest': [500000],
            'days_to_expiry': [5],
            'positioning_ratio': [0.3],
            'roll_pressure': [1.5]
        })

        calc = RollPressureCalculator(mock_config)
        result = calc.add_alert_column(data)

        assert len(result) == 1

    def test_very_high_positioning(self, mock_config):
        """Test with extremely high positioning ratio."""
        data = pd.DataFrame({
            'date': [pd.Timestamp('2025-08-01')],
            'market': ['WTI'],
            'spec_net_long': [900000],  # Very high
            'open_interest': [100000],  # Low OI
            'days_to_expiry': [10]
        })

        calc = RollPressureCalculator(mock_config)

        data['days_to_expiry'] = 10
        result = calc.calculate_roll_pressure(data)

        # Should be clipped to max_value (2.0)
        assert result['roll_pressure'].iloc[0] <= 2.0


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
