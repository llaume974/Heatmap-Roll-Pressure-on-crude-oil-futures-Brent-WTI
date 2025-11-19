"""
Excel export with conditional formatting and alerts.

Creates Excel workbook with:
1. Data sheet: Full roll pressure data table
2. Heatmap sheet: Matrix view with conditional formatting
3. Summary banner: Alert count and metadata
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict
from loguru import logger


class ExcelAlertExporter:
    """Export roll pressure data to Excel with formatting and alerts."""

    def __init__(self, config: Optional[Dict] = None):
        """
        Initialize Excel exporter.

        Args:
            config: Configuration dictionary
        """
        self.config = config or {}

        # Thresholds
        thresholds_config = self.config.get('thresholds', {})
        self.green_max = thresholds_config.get('green_max', 0.40)
        self.orange_max = thresholds_config.get('orange_max', 0.65)

        # Output path
        output_config = self.config.get('output_files', {})
        self.excel_path = output_config.get('excel', 'output/roll_pressure_latest.xlsx')

    def create_data_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create data sheet with full table.

        Args:
            wb: Workbook object
            df: Roll pressure DataFrame
        """
        logger.info("Creating data sheet...")

        ws = wb.create_sheet("Data", 0)

        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Header formatting
                if r_idx == 1:
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add conditional formatting for roll_pressure column
        if 'roll_pressure' in df.columns:
            rp_col_idx = df.columns.get_loc('roll_pressure') + 1
            rp_col_letter = ws.cell(row=1, column=rp_col_idx).column_letter

            # Green rule (< green_max)
            green_rule = CellIsRule(
                operator='lessThan',
                formula=[str(self.green_max)],
                fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                font=Font(color='006100')
            )

            # Orange rule (between green_max and orange_max)
            orange_rule = CellIsRule(
                operator='between',
                formula=[str(self.green_max), str(self.orange_max)],
                fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                font=Font(color='9C6500')
            )

            # Red rule (> orange_max)
            red_rule = CellIsRule(
                operator='greaterThan',
                formula=[str(self.orange_max)],
                fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                font=Font(color='9C0006')
            )

            # Apply rules
            range_str = f"{rp_col_letter}2:{rp_col_letter}{len(df) + 1}"
            ws.conditional_formatting.add(range_str, green_rule)
            ws.conditional_formatting.add(range_str, orange_rule)
            ws.conditional_formatting.add(range_str, red_rule)

        # Highlight ALERTE_48H column
        if 'ALERTE_48H' in df.columns:
            alert_col_idx = df.columns.get_loc('ALERTE_48H') + 1
            alert_col_letter = ws.cell(row=1, column=alert_col_idx).column_letter

            # True = Red background
            alert_rule = CellIsRule(
                operator='equal',
                formula=['TRUE'],
                fill=PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
                font=Font(color='FFFFFF', bold=True)
            )

            range_str = f"{alert_col_letter}2:{alert_col_letter}{len(df) + 1}"
            ws.conditional_formatting.add(range_str, alert_rule)

        logger.info(f"✓ Data sheet created with {len(df)} rows")

    def create_heatmap_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create heatmap sheet with matrix view and conditional formatting.

        Args:
            wb: Workbook object
            df: Roll pressure DataFrame
        """
        logger.info("Creating heatmap sheet...")

        ws = wb.create_sheet("Heatmap")

        # Pivot data for heatmap
        pivot = df.pivot_table(
            index='market',
            columns='date',
            values='roll_pressure',
            aggfunc='first'
        )

        # Write pivot table
        for r_idx, row in enumerate(dataframe_to_rows(pivot, index=True, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Format dates
                if r_idx == 1 and c_idx > 1:  # Header dates
                    if isinstance(value, datetime):
                        cell.value = value.strftime('%Y-%m-%d')
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = Alignment(horizontal='center', textRotation=45)

                # Format market names
                if c_idx == 1 and r_idx > 1:  # Market names
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='right')

                # Format numbers
                if r_idx > 1 and c_idx > 1 and isinstance(value, (int, float)):
                    cell.number_format = '0.000'

        # Apply conditional formatting to data cells
        if len(pivot) > 0:
            # Determine range (excluding headers)
            max_row = len(pivot) + 1
            max_col = len(pivot.columns) + 1

            # Green
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            green_rule = CellIsRule(
                operator='lessThan',
                formula=[str(self.green_max)],
                fill=green_fill,
                font=Font(color='006100')
            )

            # Orange
            orange_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            orange_rule = CellIsRule(
                operator='between',
                formula=[str(self.green_max), str(self.orange_max)],
                fill=orange_fill,
                font=Font(color='9C6500')
            )

            # Red
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            red_rule = CellIsRule(
                operator='greaterThan',
                formula=[str(self.orange_max)],
                fill=red_fill,
                font=Font(color='9C0006', bold=True)
            )

            # Apply to range
            range_str = f"B2:{ws.cell(row=max_row, column=max_col).coordinate}"
            ws.conditional_formatting.add(range_str, green_rule)
            ws.conditional_formatting.add(range_str, orange_rule)
            ws.conditional_formatting.add(range_str, red_rule)

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, len(pivot.columns) + 2):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12

        logger.info(f"✓ Heatmap sheet created: {pivot.shape[0]} markets × {pivot.shape[1]} dates")

    def create_summary_banner(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create summary sheet with alert information.

        Args:
            wb: Workbook object
            df: Roll pressure DataFrame
        """
        logger.info("Creating summary banner...")

        ws = wb.create_sheet("Summary", 0)

        # Calculate metrics
        total_rows = len(df)
        alert_count = df['ALERTE_48H'].sum() if 'ALERTE_48H' in df.columns else 0
        markets = df['market'].unique().tolist() if len(df) > 0 else []

        # Handle empty DataFrame
        if len(df) > 0 and 'date' in df.columns:
            date_min = df['date'].min()
            date_max = df['date'].max()
            if pd.notna(date_min) and pd.notna(date_max):
                date_range = f"{date_min.strftime('%Y-%m-%d')} to {date_max.strftime('%Y-%m-%d')}"
            else:
                date_range = "No data"
        else:
            date_range = "No data"

        # Get latest alerts
        latest_alerts = df[df['ALERTE_48H'] == True].sort_values('date', ascending=False) if alert_count > 0 else pd.DataFrame()

        # Title
        ws['A1'] = 'ROLL PRESSURE ALERT REPORT'
        ws['A1'].font = Font(size=18, bold=True, color='1F4E78')

        # Generation date
        ws['A3'] = 'Generated:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A3'].font = Font(bold=True)

        # Date range
        ws['A4'] = 'Date Range:'
        ws['B4'] = date_range
        ws['A4'].font = Font(bold=True)

        # Markets
        ws['A5'] = 'Markets:'
        ws['B5'] = ', '.join(markets)
        ws['A5'].font = Font(bold=True)

        # Total records
        ws['A6'] = 'Total Records:'
        ws['B6'] = total_rows
        ws['A6'].font = Font(bold=True)

        # Alert count
        ws['A8'] = 'ACTIVE ALERTS:'
        ws['B8'] = alert_count
        ws['A8'].font = Font(size=14, bold=True, color='C00000')
        ws['B8'].font = Font(size=14, bold=True, color='C00000')
        ws['B8'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Alert details
        if alert_count > 0:
            ws['A10'] = 'LATEST ALERTS:'
            ws['A10'].font = Font(size=12, bold=True)

            # Headers
            headers = ['Date', 'Market', 'Roll Pressure', 'Days to Expiry']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=11, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

            # Alert rows
            for row_idx, (_, alert_row) in enumerate(latest_alerts.head(10).iterrows(), 12):
                ws.cell(row=row_idx, column=1, value=alert_row['date'].strftime('%Y-%m-%d'))
                ws.cell(row=row_idx, column=2, value=alert_row['market'])
                ws.cell(row=row_idx, column=3, value=float(alert_row['roll_pressure']))
                ws.cell(row=row_idx, column=4, value=int(alert_row['days_to_expiry']))

                # Highlight row
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
                    )

        else:
            ws['A10'] = 'No active alerts.'
            ws['A10'].font = Font(size=12, color='006100')

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

        logger.info(f"✓ Summary banner created ({alert_count} alerts)")

    def export_to_excel(self, df: pd.DataFrame, output_path: Optional[str] = None) -> str:
        """
        Export roll pressure data to Excel with all sheets.

        Args:
            df: Roll pressure DataFrame
            output_path: Custom output path (optional)

        Returns:
            Path to saved Excel file
        """
        if output_path is None:
            output_path = self.excel_path

        logger.info(f"Exporting to Excel: {output_path}")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create sheets
        self.create_summary_banner(wb, df)
        self.create_data_sheet(wb, df)
        self.create_heatmap_sheet(wb, df)

        # Save
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"✓ Excel file saved to {output_path}")

        return str(output_path)


def export_to_excel(df: pd.DataFrame, config: Optional[Dict] = None, output_path: Optional[str] = None) -> str:
    """
    Convenience function to export to Excel.

    Args:
        df: Roll pressure DataFrame
        config: Configuration dict
        output_path: Custom output path

    Returns:
        Path to saved file
    """
    exporter = ExcelAlertExporter(config)
    return exporter.export_to_excel(df, output_path)
